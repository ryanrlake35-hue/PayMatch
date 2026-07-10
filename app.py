"""
PayMatch by ProPayHR — Benefits Reconciliation Platform
"""
import base64, io, re, warnings, os, threading, time
from datetime import datetime
from dotenv import load_dotenv
from supabase import create_client, Client as SupabaseClient
from pydantic import BaseModel, field_validator

load_dotenv()

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

warnings.filterwarnings("ignore")

# ── EXCEL CONSTANTS ───────────────────────────────────────────────
NAVY = "1F3864"; WHITE = "FFFFFF"; SEC = "D9E1F2"; TF = "1F3864"
THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AR   = "Arial"
AF   = {"OK":"E2EFDA","Add":"DDEBF7","Change":"FCE4D6","Review":"FFF2CC"}

DISPLAY_COLS = ["Location","First Name","Last Name","Status","Benefit",
                "Broker /pay","Paycor /pay","Difference /pay","Action","Note"]
_APP_DIR    = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(_APP_DIR, "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

# Broker "Benefit" value → Paycor amount column
BENEFIT_CROSSWALK = {
    "Medical":                         "Medical Amount",
    "Dental":                          "Dent Amount",
    "Vision":                          "Vis Amount",
    "Accident":                        "Acc Amount",
    "Critical Illness":                "CritIll Amount",
    "Hospital Indemnity":              "HospIND Amount",
    "Identity Protection":             "ID Protect Amount",
    "Legal":                           "Legal Amount",
    "Voluntary Life":                  "Life/AD&D Amount",
    "Voluntary Long-Term Disability":  "LongTermD Amount",
    "Voluntary Short-Term Disability": "ShortTermD Amount",
}

PAYCOR_BENEFIT_COLS = [
    "Medical Amount","MedicalER2 Amount","Vis Amount","Dent Amount",
    "Life/AD&D Amount","LongTermD Amount","ShortTermD Amount","CritIll Amount",
    "Acc Amount","HospIND Amount","ID Protect Amount","Legal Amount",
]
# Employer-share columns: never part of employee-cost matching
PAYCOR_EXCLUDED_COLS = {"MedicalER2 Amount"}

# Broker division name → Paycor Client Name (both normalized lowercase)
DIVISION_ALIASES = {
    "lotus therapy": "lotus therapy partners llc",
}

def _money(v):
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:    return float(s)
    except: return 0.0

def _norm_name(s):
    return re.sub(r"[^a-z]", "", str(s).lower())

def _norm_div(s, apply_alias=False):
    d = re.sub(r"\s+", " ", str(s).strip().lower())
    return DIVISION_ALIASES.get(d, d) if apply_alias else d

def _split_paycor_full_name(full):
    # Paycor Full Name is "Last, First Middle"
    parts = str(full).split(",", 1)
    last  = parts[0].strip()
    rest  = parts[1].strip().split() if len(parts) > 1 else []
    first = rest[0] if rest else ""
    return first, last

# ── DESIGN SYSTEM CSS ─────────────────────────────────────────────
APP_CSS = """
@import url('https://fonts.googleapis.com/css2?family=Fraunces:ital,opsz,wght@0,9..144,300;0,9..144,400;1,9..144,300;1,9..144,400&family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

/* ── Brand tokens ─────────────────────────────────── */
:root {
    --ink:            #2A0E14;
    --stone:          #F1ECE6;
    --paper:          #FFFFFF;
    --burgundy:       #8B1A2F;
    --burgundy-dark:  #6B1322;
    --gold:           #C9A227;
    --gold-muted:     rgba(201,162,39,0.14);
    --gold-border:    rgba(201,162,39,0.30);
    --green:          #1E7A4C;
    --green-bg:       #E3F0E8;
    --red:            #B23A22;
    --red-bg:         #F8E6E0;
    --text-soft:      #5C544E;
    --border:         #E3E0D8;
    --border-strong:  #CCC7C0;
}

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}
#MainMenu, footer, header { visibility: hidden; }
.stApp { background: var(--stone) !important; }
.block-container {
    padding-top: 0 !important;
    padding-bottom: 2.5rem !important;
    max-width: 1140px;
}

/* ── HERO HEADER ──────────────────────────────────── */
.pm-hero {
    background: var(--ink);
    padding: 1.9rem 2.5rem 1.6rem;
    margin: -1rem -1rem 0;
    position: relative;
    overflow: hidden;
}
.pm-hero::before {
    content: '';
    position: absolute;
    top: -80px; right: -80px;
    width: 360px; height: 360px;
    background: radial-gradient(circle, rgba(201,162,39,0.08) 0%, transparent 65%);
    pointer-events: none;
}
.pm-hero-inner {
    position: relative;
    z-index: 1;
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    gap: 1.5rem;
    flex-wrap: wrap;
}
.pm-hero-left { flex: 1; min-width: 240px; }
.pm-superlabel {
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 0.22em;
    text-transform: uppercase;
    color: rgba(255,255,255,0.35);
    margin-bottom: 0.6rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
.pm-superlabel::before {
    content: '';
    display: inline-block;
    width: 18px;
    height: 1.5px;
    background: var(--gold);
    border-radius: 2px;
}
.pm-wordmark {
    font-family: 'Fraunces', serif;
    font-size: 2.3rem;
    font-weight: 300;
    color: #FFFFFF;
    letter-spacing: -0.04em;
    line-height: 1;
    margin-bottom: 0.3rem;
}
.pm-wordmark em {
    color: var(--gold);
    font-style: italic;
}
.pm-descriptor {
    font-size: 0.82rem;
    font-weight: 400;
    color: rgba(255,255,255,0.42);
    letter-spacing: 0.01em;
    line-height: 1.5;
    margin-top: 0.35rem;
}
.pm-chips {
    display: flex;
    flex-wrap: wrap;
    gap: 0.45rem;
    margin-top: 1.25rem;
}
.pm-chip {
    font-size: 0.68rem;
    font-weight: 500;
    color: rgba(255,255,255,0.5);
    background: rgba(255,255,255,0.055);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 999px;
    padding: 0.22rem 0.8rem;
    letter-spacing: 0.01em;
}
.pm-chip-check { color: var(--gold); margin-right: 0.25rem; }

/* ── GOLD DIVIDER ─────────────────────────────────── */
.gold-rule {
    height: 4px;
    background: var(--gold);
    border-radius: 0;
    margin: 0 0 1.5rem;
    opacity: 0.9;
    animation: pm-pulse 4s ease-in-out infinite;
}

/* ── SECTION LABELS ───────────────────────────────── */
.slabel {
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 0.2em;
    text-transform: uppercase;
    color: var(--burgundy);
    margin-bottom: 0.5rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
.slabel::before {
    content: '';
    display: inline-block;
    width: 18px;
    height: 2px;
    background: var(--gold);
    border-radius: 2px;
    flex-shrink: 0;
}
.sdesc {
    font-size: 0.84rem;
    color: var(--text-soft);
    line-height: 1.65;
    margin-bottom: 1rem;
}

/* ── CARDS ────────────────────────────────────────── */
.card {
    background: var(--paper);
    border: 1px solid var(--border);
    border-radius: 13px;
    padding: 1.3rem 1.5rem;
    margin-bottom: 0.85rem;
}
.card-title {
    font-family: 'Fraunces', serif;
    font-size: 0.98rem;
    font-weight: 400;
    color: var(--ink);
    display: flex;
    align-items: center;
    gap: 0.6rem;
    margin-bottom: 0.3rem;
}
.step-badge {
    width: 25px;
    height: 25px;
    background: var(--burgundy);
    color: white;
    border-radius: 7px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-size: 0.7rem;
    font-weight: 700;
    font-family: 'Inter', sans-serif;
    flex-shrink: 0;
}
.file-tag {
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: var(--gold);
    background: var(--gold-muted);
    border: 1px solid var(--gold-border);
    border-radius: 4px;
    padding: 0.1rem 0.48rem;
}
.card-desc {
    font-size: 0.81rem;
    color: var(--text-soft);
    line-height: 1.6;
    padding-left: 2.05rem;
}

/* ── CARD HOVER + STEP CARD 3D TILT ──────────────── */
.card { transition: transform 150ms ease; }
.card:not(.pm-step-card):hover { transform: translateY(-1px); }
.pm-step-card {
    transition: transform 200ms ease;
    transform-style: preserve-3d;
    will-change: transform;
}
.pm-step-card:hover {
    transform: perspective(600px) rotateY(4deg) rotateX(-2deg) translateY(-3px);
}

/* ── BUTTONS ──────────────────────────────────────── */
.stButton > button {
    background: var(--burgundy) !important;
    color: white !important;
    border: 1.5px solid var(--burgundy) !important;
    border-radius: 999px !important;
    padding: 0.72rem 2rem !important;
    font-size: 0.9rem !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    letter-spacing: 0.015em !important;
    box-shadow: none !important;
    transition: all 0.15s ease !important;
    width: 100% !important;
    min-height: 44px !important;
}
.stButton > button:hover:not(:disabled) {
    background: var(--burgundy-dark) !important;
    border-color: var(--burgundy-dark) !important;
    transform: translateY(-1px) !important;
}
.stButton > button:active:not(:disabled) {
    transform: scale(0.97) !important;
}
.stButton > button:disabled {
    opacity: 0.35 !important;
    cursor: not-allowed !important;
}

/* Download buttons — gold */
[data-testid="stDownloadButton"] > button {
    background: var(--gold) !important;
    color: white !important;
    border: 1.5px solid var(--gold) !important;
    border-radius: 999px !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    box-shadow: none !important;
    width: 100% !important;
    padding: 0.72rem !important;
    font-size: 0.88rem !important;
    letter-spacing: 0.01em !important;
    transition: all 0.15s ease !important;
    min-height: 44px !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #a88a1f !important;
    border-color: #a88a1f !important;
    transform: translateY(-1px) !important;
}

/* ── INPUTS ───────────────────────────────────────── */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stSelectbox > div > div > div,
.stDateInput > div > div > input {
    background-color: var(--paper) !important;
    color: var(--ink) !important;
    border-color: var(--border-strong) !important;
    border-radius: 8px !important;
    font-size: 0.88rem !important;
    font-family: 'Inter', sans-serif !important;
    padding: 0.52rem 0.75rem !important;
}
.stTextInput > div > div > input::placeholder,
.stNumberInput > div > div > input::placeholder {
    color: #A89E96 !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus {
    border-color: var(--burgundy) !important;
    box-shadow: 0 0 0 3px rgba(139,26,47,0.1) !important;
    outline: none !important;
}
.stSelectbox > div > div {
    background: var(--paper) !important;
    color: var(--ink) !important;
}
label, .stTextInput label, .stNumberInput label, .stSelectbox label {
    color: var(--ink) !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
}

/* ── METRIC TILES ─────────────────────────────────── */
.pm-metrics {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 0.85rem;
    margin: 1rem 0;
}
.pm-metric {
    background: var(--paper);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1rem 1.2rem;
    border-top: 3px solid var(--burgundy);
}
.pm-metric-label {
    font-size: 0.64rem;
    font-weight: 700;
    color: var(--text-soft);
    text-transform: uppercase;
    letter-spacing: 0.13em;
    font-family: 'Inter', sans-serif;
    margin-bottom: 0.3rem;
}
.pm-metric-value {
    font-size: 1.8rem;
    font-weight: 500;
    color: var(--ink);
    letter-spacing: -0.03em;
    line-height: 1.1;
    font-family: 'IBM Plex Mono', monospace;
}
@keyframes pm-pop {
    0%   { opacity: 0; transform: scale(0.82); }
    60%  { opacity: 1; transform: scale(1.05); }
    100% { opacity: 1; transform: scale(1); }
}
.pm-countup {
    display: inline-block;
    font-variant-numeric: tabular-nums;
    animation: pm-pop 350ms ease-out both;
}

/* ── STATUS BADGES (results table legend) ─────────── */
.badge {
    display: inline-flex;
    align-items: center;
    border-radius: 999px;
    padding: 0.18rem 0.65rem;
    font-size: 0.72rem;
    font-weight: 600;
    font-family: 'Inter', sans-serif;
    white-space: nowrap;
}
.badge-ok     { background: var(--green-bg); color: var(--green); }
.badge-add    { background: rgba(139,26,47,0.1); color: var(--burgundy); }
.badge-change { background: var(--gold-muted); color: #7a5a00; }
.badge-review { background: var(--red-bg); color: var(--red); }

/* ── RESULT BAR ───────────────────────────────────── */
.res-bar {
    background: var(--paper);
    border: 1px solid var(--border);
    border-left: 4px solid var(--burgundy);
    border-radius: 13px;
    padding: 1.05rem 1.35rem;
    margin: 1rem 0;
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 0.5rem;
}
.res-bar-ttl {
    font-family: 'Fraunces', serif;
    font-size: 1rem;
    font-weight: 400;
    color: var(--ink);
}
.res-bar-sub { font-size: 0.78rem; color: var(--text-soft); margin-top: 0.1rem; }
.run-ts { font-size: 0.72rem; color: #9D8E88; margin-top: 0.3rem; display: flex; align-items: center; gap: 0.3rem; }
.res-bar-count {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.2rem;
    font-weight: 500;
    color: var(--burgundy);
}

/* ── URGENCY BOX ──────────────────────────────────── */
.urgency {
    background: var(--red-bg);
    border: 1px solid rgba(178,58,34,0.18);
    border-left: 4px solid var(--red);
    border-radius: 13px;
    padding: 1rem 1.35rem;
    margin: 0.85rem 0;
    display: flex;
    align-items: flex-start;
    gap: 1rem;
}
.urgency-amount {
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 500;
    color: var(--red);
    font-size: 1rem;
}
.urgency-note { color: #7F1D1D; font-size: 0.82rem; margin-top: 0.22rem; line-height: 1.55; }

/* ── INFO BOX ─────────────────────────────────────── */
.info-box {
    background: var(--gold-muted);
    border: 1px solid var(--gold-border);
    border-radius: 10px;
    padding: 0.9rem 1.1rem;
    margin-top: 0.85rem;
    font-size: 0.82rem;
    color: var(--text-soft);
    line-height: 1.65;
}
.info-box b { color: var(--ink); display: block; margin-bottom: 0.3rem; font-weight: 700; }

/* ── EMPTY STATE ──────────────────────────────────── */
.empty-prompt {
    text-align: center;
    color: #9D8E88;
    font-size: 0.88rem;
    padding: 2.5rem 1.5rem;
    background: var(--paper);
    border-radius: 13px;
    border: 1.5px dashed var(--border-strong);
    margin-top: 0.5rem;
    line-height: 1.65;
}

/* ── DOWNLOAD SECTION ─────────────────────────────── */
.dl-section {
    background: var(--paper);
    border: 1px solid var(--border);
    border-top: 3px solid var(--gold);
    border-radius: 13px;
    padding: 1.15rem 1.4rem;
    margin-top: 1.1rem;
}
.dl-title {
    font-family: 'Fraunces', serif;
    font-size: 0.95rem;
    font-weight: 400;
    color: var(--ink);
    margin-bottom: 0.75rem;
}
.dl-desc { font-size: 0.8rem; color: var(--text-soft); margin-bottom: 0.65rem; line-height: 1.55; }

/* ── TREND CARD ───────────────────────────────────── */
.trend-card {
    background: var(--paper);
    border: 1px solid var(--border);
    border-radius: 13px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}
.delta { border-radius: 10px; padding: 0.85rem 1.35rem; font-weight: 600; font-size: 0.84rem; margin-top: 0.6rem; }
.delta-dn { background: var(--green-bg); border: 1px solid rgba(30,122,76,0.2); color: var(--green); }
.delta-up { background: var(--red-bg);   border: 1px solid rgba(178,58,34,0.2); color: var(--red); }

/* ── NOTES ────────────────────────────────────────── */
.notes-wrap {
    background: var(--paper);
    border: 1px solid var(--border);
    border-radius: 13px;
    padding: 1.35rem 1.5rem;
    margin-top: 1rem;
}
.notes-label {
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--burgundy);
    margin-bottom: 0.5rem;
}
/* ── FOOTER ───────────────────────────────────────── */
.pm-footer {
    text-align: center;
    color: #9D8E88;
    font-size: 0.72rem;
    padding: 1.5rem 1rem 0.5rem;
    margin-top: 1.5rem;
    letter-spacing: 0.02em;
    border-top: 1px solid var(--border);
}
.pm-footer strong { color: var(--burgundy); font-weight: 600; }

/* ── MISC ─────────────────────────────────────────── */
.stAlert { border-radius: 10px !important; font-size: 0.86rem !important; }
hr { border-color: var(--border) !important; margin: 1rem 0 !important; }
[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    overflow: hidden !important;
    border: 1px solid var(--border) !important;
    box-shadow: none !important;
}
.stSpinner > div { border-top-color: var(--burgundy) !important; }

/* ── ANIMATIONS ───────────────────────────────────── */
@keyframes pm-pulse {
    0%, 100% { opacity: 0.9; }
    50%       { opacity: 0.6; }
}

/* ── MOBILE ───────────────────────────────────────── */
@media (max-width: 768px) {
    .block-container { padding: 0.4rem 0.4rem 2rem !important; }
    .pm-hero { margin: -0.5rem -0.5rem 0; padding: 1.2rem 1.1rem 1rem; }
    .pm-wordmark { font-size: 1.7rem !important; }
    .pm-chips { display: none !important; }
    .stButton > button { min-height: 50px !important; font-size: 0.95rem !important; }
    [data-testid="stDownloadButton"] > button { min-height: 50px !important; }
    .pm-metrics { grid-template-columns: repeat(3, 1fr); }
    .pm-metric-value { font-size: 1.4rem; }
}

/* ── REDUCED MOTION ───────────────────────────────── */
@media (prefers-reduced-motion: reduce) {
    .gold-rule                           { animation: none !important; }
    .pm-countup                          { animation: none !important; }
    .pm-step-card, .pm-step-card:hover   { transition: none !important; transform: none !important; }
    .card:not(.pm-step-card)             { transition: none !important; }
    .stButton > button,
    [data-testid="stDownloadButton"] > button { transition: none !important; }
}
"""

# ── SUPABASE ──────────────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")

@st.cache_resource
def _sb() -> SupabaseClient:
    return create_client(SUPABASE_URL, SUPABASE_KEY)

# ── PYDANTIC MODEL ────────────────────────────────────────────────
class ReconciliationRecord(BaseModel):
    client: str
    period: str
    run_by: str
    run_date: str
    run_time: str
    total_lines: int
    ok_count: int
    add_count: int
    change_count: int
    review_count: int
    discrepancies: int
    monthly_at_stake: float
    status: str = "Incomplete"
    report_file: str = ""
    notes: str = ""

    @field_validator("client", "period")
    @classmethod
    def not_empty(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("Field cannot be empty")
        return v.strip()

    @field_validator("total_lines", "ok_count", "add_count", "change_count", "review_count", "discrepancies")
    @classmethod
    def non_negative(cls, v: int) -> int:
        if v < 0:
            raise ValueError("Count must be >= 0")
        return v

    @field_validator("monthly_at_stake")
    @classmethod
    def valid_amount(cls, v: float) -> float:
        if v < 0:
            raise ValueError("monthly_at_stake must be >= 0")
        return round(v, 2)

# ── FILE PARSERS ──────────────────────────────────────────────────
def read_file(file):
    n = file.name.lower()
    if n.endswith(".csv"):   return pd.read_csv(file, dtype=str)
    elif n.endswith(".xls"): return pd.read_excel(file, dtype=str, engine="xlrd")
    else:                    return pd.read_excel(file, dtype=str)

def parse_paycor(file):
    df = read_file(file)
    df.columns = [str(c).strip() for c in df.columns]
    # Drop the export footer ("Count: N") and any row without a Client ID
    cid = df.get("Client ID", pd.Series(dtype=str)).astype(str).str.strip()
    df = df[(cid != "") & (cid.str.lower() != "nan")].reset_index(drop=True)
    ben_cols = [c for c in PAYCOR_BENEFIT_COLS if c in df.columns]
    for col in ben_cols:
        df[col] = df[col].map(lambda v: _money(v) if str(v).strip() not in ("", "nan", "None") else None)
    id_cols = [c for c in ["Client Name","Employee Number","Full Name","Status Type"] if c in df.columns]
    long = df.melt(id_vars=id_cols, value_vars=ben_cols,
                   var_name="Paycor Column", value_name="Paycor Amount").dropna(subset=["Paycor Amount"])
    # Split checks: one line per employee per benefit
    long = (long.groupby(["Client Name","Employee Number","Full Name","Paycor Column"], as_index=False)
                .agg({"Paycor Amount":"sum","Status Type":"first"}))
    return long

def parse_broker_and_crosswalk(file):
    df = read_file(file)
    df.columns = [str(c).strip() for c in df.columns]
    for col in ("EE Cost", "ER Cost", "Benefit Amount"):
        if col in df.columns:
            df[col] = df[col].map(_money)
    return df.reset_index(drop=True), dict(BENEFIT_CROSSWALK)

# ── RECONCILIATION ENGINE ─────────────────────────────────────────
def reconcile(broker_df, paycor_df, crosswalk, tolerance):
    rev_cw = {v: k for k, v in crosswalk.items()}
    # One entry per Paycor person, plus their summed per-benefit amounts
    people, amounts = {}, {}
    for _, r in paycor_df.iterrows():
        key = (str(r["Client Name"]), str(r["Employee Number"]))
        if key not in people:
            first, last = _split_paycor_full_name(r["Full Name"])
            people[key] = {"nfirst": _norm_name(first), "nlast": _norm_name(last),
                           "ndiv": _norm_div(r["Client Name"]),
                           "first": first, "last": last,
                           "loc": str(r["Client Name"]), "status": str(r.get("Status Type",""))}
            amounts[key] = {}
        amounts[key][r["Paycor Column"]] = float(r["Paycor Amount"])

    def match_person(nfirst, nlast, ndiv):
        """Returns (key, 'matched') | (None, 'unmatched') | (None, 'ambiguous')."""
        cands = [k for k, p in people.items() if p["nlast"] == nlast and p["ndiv"] == ndiv]
        pref  = [k for k in cands if nfirst and people[k]["nfirst"] and
                 (people[k]["nfirst"].startswith(nfirst) or nfirst.startswith(people[k]["nfirst"]))]
        if len(pref) == 1: return pref[0], "matched"
        if len(pref) > 1:  return None, "ambiguous"
        # Fallback: exact full-name match across all divisions, only if unique
        exact = [k for k, p in people.items() if p["nlast"] == nlast and p["nfirst"] == nfirst and nfirst]
        if len(exact) == 1: return exact[0], "matched"
        if len(exact) > 1:  return None, "ambiguous"
        return None, "unmatched"

    rows, consumed = [], set()
    match_cache = {}
    for _, emp in broker_df.iterrows():
        fn = str(emp.get("First Name","")).strip(); ln = str(emp.get("Last Name","")).strip()
        div = str(emp.get("Division","")).strip(); status = str(emp.get("Employment Status","")).strip()
        if not fn and not ln: continue
        pkey = (_norm_name(fn), _norm_name(ln), _norm_div(div, apply_alias=True))
        if pkey not in match_cache:
            match_cache[pkey] = match_person(pkey[0], pkey[1], pkey[2])
        key, mstatus = match_cache[pkey]

        benefit = str(emp.get("Benefit","")).strip()
        bval = float(emp.get("EE Cost", 0) or 0)
        pcol = crosswalk.get(benefit)
        pval = 0.0; action = ""; note = ""
        if mstatus == "unmatched":
            action = "Review - Not in Paycor"; note = "Employee not found in Paycor — possible new hire or name mismatch"
        elif mstatus == "ambiguous":
            action = "Review - Multiple matches"; note = "More than one Paycor employee matches this name — verify manually"
        elif pcol is None:
            action = "Review - Not in Paycor export"; note = "No Paycor column mapped — verify manually"
        else:
            consumed.add((key, pcol))
            amt = amounts[key].get(pcol)
            present = amt is not None and round(amt, 2) != 0
            pval = float(amt) if present else 0.0
            if not present and bval == 0:
                action = "OK"; note = ""
            elif not present:
                action = "Add - Start deduction"; note = "Enrolled with broker but zero deduction in Paycor"
            elif bval == 0:
                action = "Review - Deduction vs $0 enrollment"; note = f"Paycor withholds ${pval:.2f} but broker shows $0.00 enrollment"
            elif abs(round(bval - pval, 2)) <= tolerance:
                action = "OK"; note = ""
            else:
                action = "Change - Amount differs"; note = f"Broker ${bval:.2f} vs Paycor ${pval:.2f} per pay"
        rows.append({"Location":div,"First Name":fn,"Last Name":ln,"Status":status,
            "Benefit":benefit,"Broker /pay":round(bval,2),"Paycor /pay":round(pval,2),
            "Difference /pay":round(bval-pval,2),"Action":action,"Note":note})

    # Paycor deductions with no broker enrollment line at all
    for key, cols in amounts.items():
        p = people[key]
        for pcol, amt in cols.items():
            if pcol in PAYCOR_EXCLUDED_COLS or (key, pcol) in consumed or round(amt, 2) == 0:
                continue
            benefit = rev_cw.get(pcol, pcol)
            rows.append({"Location":p["loc"],"First Name":p["first"],"Last Name":p["last"],
                "Status":p["status"],"Benefit":benefit,"Broker /pay":0.0,"Paycor /pay":round(amt,2),
                "Difference /pay":round(-amt,2),"Action":"Review - Not in broker file",
                "Note":f"Paycor withholds ${amt:.2f} but no broker enrollment line exists"})
    return pd.DataFrame(rows).reset_index(drop=True)

# ── EXCEL BUILDER ─────────────────────────────────────────────────
def _sw(ws,wmap):
    for ci,w in wmap.items(): ws.column_dimensions[get_column_letter(ci)].width=w

def _nh(ws,txt,n):
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    c=ws["A1"]; c.value=txt; c.font=Font(name=AR,size=12,bold=True,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=22

def _kh(ws,cols,r=2):
    for h,cs,ce in cols:
        if cs!=ce: ws.merge_cells(start_row=r,start_column=cs,end_row=r,end_column=ce)
        c=ws.cell(r,cs,h); c.font=Font(name=AR,size=10,bold=True,color=WHITE)
        c.fill=PatternFill("solid",fgColor="404040"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORD
    ws.row_dimensions[r].height=18

def _kr(ws,ri,fc,label,count,meaning,action,me,as_,ae):
    ws.row_dimensions[ri].height=28
    c1=ws.cell(ri,1,label); c1.font=Font(name=AR,size=10,bold=True); c1.fill=PatternFill("solid",fgColor=fc); c1.border=BORD; c1.alignment=Alignment(vertical="center")
    c2=ws.cell(ri,2,count); c2.font=Font(name=AR,size=14,bold=True,color=NAVY); c2.fill=PatternFill("solid",fgColor=fc); c2.border=BORD; c2.alignment=Alignment(horizontal="center",vertical="center")
    ws.merge_cells(start_row=ri,start_column=3,end_row=ri,end_column=me)
    c3=ws.cell(ri,3,meaning); c3.font=Font(name=AR,size=10); c3.fill=PatternFill("solid",fgColor=fc); c3.border=BORD; c3.alignment=Alignment(vertical="center",wrap_text=True)
    ws.merge_cells(start_row=ri,start_column=as_,end_row=ri,end_column=ae)
    c4=ws.cell(ri,as_,f"Action: {action}"); c4.font=Font(name=AR,size=10,bold=True); c4.fill=PatternFill("solid",fgColor=fc); c4.border=BORD; c4.alignment=Alignment(vertical="center",wrap_text=True)

def _kt(ws,ri,lbl,cnt,summary,right,ss,se,as_,ae):
    ws.row_dimensions[ri].height=24
    for ci,v,_ in [(1,lbl,None),(2,cnt,None)]:
        c=ws.cell(ri,ci,v); c.font=Font(name=AR,size=10 if ci==1 else 14,bold=True,color=WHITE)
        c.fill=PatternFill("solid",fgColor=TF); c.border=BORD
        c.alignment=Alignment(indent=1 if ci==1 else 0,horizontal="center" if ci==2 else "left",vertical="center")
    ws.merge_cells(start_row=ri,start_column=ss,end_row=ri,end_column=se)
    c3=ws.cell(ri,ss,summary); c3.font=Font(name=AR,size=10,bold=True,color=WHITE); c3.fill=PatternFill("solid",fgColor=TF); c3.border=BORD; c3.alignment=Alignment(vertical="center",indent=1)
    ws.merge_cells(start_row=ri,start_column=as_,end_row=ri,end_column=ae)
    c4=ws.cell(ri,as_,right); c4.font=Font(name=AR,size=10,bold=True,color=WHITE); c4.fill=PatternFill("solid",fgColor=TF); c4.border=BORD; c4.alignment=Alignment(vertical="center",indent=1)

def _dh(ws,heads,row):
    for ci,h in enumerate(heads,1):
        c=ws.cell(row,ci,h); c.font=Font(name=AR,size=10,bold=True,color=WHITE)
        c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=BORD
    ws.row_dimensions[row].height=28; ws.freeze_panes=f"A{row+1}"

def _dr(ws,src,start):
    for ri,row in src[DISPLAY_COLS].iterrows():
        rn=ri+start; fc=AF.get(str(row["Action"]).split(" - ")[0],"FFFFFF")
        for ci,v in enumerate(row.values,1):
            c=ws.cell(rn,ci,v); c.font=Font(name=AR,size=10); c.fill=PatternFill("solid",fgColor=fc); c.border=BORD
            if ci in (6,7,8): c.number_format="$#,##0.00"

def _nt(ws,r,c1,lbl,c2,val,fmt="0",span=None):
    if span: ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=span)
    cx=ws.cell(r,c1,lbl); cx.font=Font(name=AR,size=10,bold=True,color=WHITE); cx.fill=PatternFill("solid",fgColor=TF); cx.border=BORD; cx.alignment=Alignment(indent=1)
    vx=ws.cell(r,c2,val); vx.font=Font(name=AR,size=10,bold=True,color=WHITE); vx.fill=PatternFill("solid",fgColor=TF); vx.border=BORD; vx.alignment=Alignment(horizontal="right"); vx.number_format=fmt

def build_action_items_only(df, client_name="Client"):
    disc=df[df["Action"]!="OK"].reset_index(drop=True)
    ct={"Add":int(df["Action"].str.startswith("Add").sum()),"Change":int(df["Action"].str.startswith("Change").sum()),"Review":int(df["Action"].str.startswith("Review").sum())}
    tdisc=ct["Add"]+ct["Change"]+ct["Review"]; n_emps=disc["First Name"].nunique()
    add_mo=round(df[df["Action"].str.startswith("Add")]["Broker /pay"].sum()*26/12,2)
    wb=Workbook(); ws=wb.active; ws.title="Action Items"; ws.sheet_view.showGridLines=False
    _sw(ws,{1:18,2:10,3:26,4:46,5:12,6:12,7:11,8:28,9:40})
    _nh(ws,f"PayMatch  |  {client_name}  |  Action Items  |  {datetime.now().strftime('%B %d, %Y')}",9)
    ws.merge_cells("A2:I2"); c=ws["A2"]
    c.value=f"Total: {tdisc} lines  |  {n_emps} employees  |  Add: {ct['Add']}  |  Change: {ct['Change']}  |  Review: {ct['Review']}  |  ${add_mo:,.2f}/month not being collected"
    c.font=Font(name=AR,size=10,bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor="2d5299")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[2].height=18
    _kh(ws,[("Color / Action",1,1),("Count",2,2),("What it means",3,5),("What to do",6,9)],r=3)
    aik=[("DDEBF7","Blue  -  Add",ct["Add"],"Enrolled with broker but zero deduction in Paycor.","Set up the deduction in Paycor immediately."),
         ("FCE4D6","Pink  -  Change",ct["Change"],"Both sides have the benefit but dollar amounts differ.","Update the deduction amount in Paycor to match the broker."),
         ("FFF2CC","Yellow  -  Review",ct["Review"],"Benefit not in Paycor export, or employee name did not match.","Verify manually in Paycor.")]
    for ri,(fc,lb,cnt,mn,ac) in enumerate(aik,4): _kr(ws,ri,fc,lb,cnt,mn,ac,5,6,9)
    _kt(ws,7,"TOTAL ACTION ITEMS",tdisc,f"{ct['Add']} Add  +  {ct['Change']} Change  +  {ct['Review']} Review  =  {tdisc} lines",f"{n_emps} employees need attention",3,5,6,9)
    ws.row_dimensions[8].height=8; _dh(ws,DISPLAY_COLS,9); ws.column_dimensions["J"].width=40; _dr(ws,disc,10)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def build_excel(df, client_name="Client", period=""):
    header_label = client_name.strip() or period or "Client"
    disc=df[df["Action"]!="OK"].reset_index(drop=True)
    ct={"OK":int((df["Action"]=="OK").sum()),"Add":int(df["Action"].str.startswith("Add").sum()),
        "Change":int(df["Action"].str.startswith("Change").sum()),"Review":int(df["Action"].str.startswith("Review").sum())}
    total=len(df); tdisc=ct["Add"]+ct["Change"]+ct["Review"]
    adds=df[df["Action"].str.startswith("Add")]; chgs=df[df["Action"].str.startswith("Change")]
    add_mo=round(adds["Broker /pay"].sum()*26/12,2); cdiff=round(chgs["Difference /pay"].abs().sum(),2)
    locs=df[df["Action"]!="OK"].groupby("Location").size().sort_values(ascending=False)
    n_emps=disc["First Name"].nunique(); wb=Workbook()
    d=wb.active; d.title="Dashboard"; d.sheet_view.showGridLines=False
    for col in "ABCDEFGH": d.column_dimensions[col].width=13
    d.column_dimensions["E"].width=18
    def m(r): d.merge_cells(r)
    def p(co,v=None,font=None,fill=None,align=None,fmt=None,border=False):
        c=d[co]
        if v is not None: c.value=v
        c.font=font or Font(name=AR,size=11)
        if fill: c.fill=PatternFill("solid",fgColor=fill)
        if align: c.alignment=align
        if fmt: c.number_format=fmt
        if border: c.border=BORD
        return c
    m("A1:H1"); p("A1",f"ProPayHR  |  {header_label} Benefits Reconciliation  |  {period or datetime.now().strftime('%B %d, %Y')}",font=Font(name=AR,size=14,bold=True,color=WHITE),fill=NAVY,align=Alignment(vertical="center",indent=1)); d.row_dimensions[1].height=28
    m("A2:H2"); p("A2","Broker enrollment (per pay) vs Paycor deductions (per pay)  |  Direct per-pay comparison",font=Font(name=AR,size=9,italic=True,color="595959"))
    tiles=[("A","Lines compared",total,"0","DCE6F1"),("C","Matched (OK)",ct["OK"],"0","E2EFDA"),("E","Require action",ct["Add"]+ct["Change"],"0","FCE4D6"),("G","Monthly $ at stake",add_mo,"$#,##0.00","FFF2CC")]
    d.row_dimensions[4].height=16; d.row_dimensions[5].height=30
    for col,lab,val,fmt,fill in tiles:
        c2=chr(ord(col)+1); m(f"{col}4:{c2}4"); p(f"{col}4",lab,font=Font(name=AR,size=10,color="404040"),fill=fill,align=Alignment(horizontal="center"))
        m(f"{col}5:{c2}5"); p(f"{col}5",val,font=Font(name=AR,size=16,bold=True,color=NAVY),fill=fill,align=Alignment(horizontal="center"),fmt=fmt)
    m("A7:D7"); p("A7","Reconciliation outcome",font=Font(name=AR,size=12,bold=True,color=NAVY),fill=SEC,align=Alignment(indent=1))
    outs=[("Matched (OK)",ct["OK"],"E2EFDA"),("Add - start deduction",ct["Add"],"DDEBF7"),("Change - amount differs",ct["Change"],"FCE4D6"),("Review - manual check",ct["Review"],"FFF2CC")]
    for i,(lab,cnt,fc) in enumerate(outs):
        r=8+i; m(f"A{r}:C{r}"); cx=d[f"A{r}"]; cx.value=lab; cx.font=Font(name=AR,size=10); cx.fill=PatternFill("solid",fgColor=fc); cx.border=BORD
        vx=d[f"D{r}"]; vx.value=cnt; vx.font=Font(name=AR,size=10,bold=True); vx.fill=PatternFill("solid",fgColor=fc); vx.alignment=Alignment(horizontal="right"); vx.border=BORD
    _nt(d,12,1,"TOTAL LINES",4,total,span=3)
    for i,(lab,cnt,_) in enumerate(outs): d.cell(70+i,1,lab.split(" - ")[0]); d.cell(70+i,2,cnt)
    dg=DoughnutChart(); dg.title="Outcome"; dg.height=6.5; dg.width=9.5; dg.style=10; dg.varyColors=True
    dg.add_data(Reference(d,min_col=2,min_row=70,max_row=73)); dg.set_categories(Reference(d,min_col=1,min_row=70,max_row=73))
    dl=DataLabelList(); dl.showVal=True; dl.showSerName=False; dl.showCatName=False; dl.showPercent=False; dg.dataLabels=dl; d.add_chart(dg,"E7")
    m("A14:D14"); p("A14","Discrepancies by location",font=Font(name=AR,size=12,bold=True,color=NAVY),fill=SEC,align=Alignment(indent=1))
    m("A15:C15"); p("A15","Location",font=Font(name=AR,size=10,bold=True),fill="F2F2F2",border=True); p("D15","Issues",font=Font(name=AR,size=10,bold=True),fill="F2F2F2",align=Alignment(horizontal="right"),border=True)
    for i,(loc,cnt) in enumerate(locs.items()):
        r=16+i; m(f"A{r}:C{r}"); d[f"A{r}"].value=loc; d[f"A{r}"].font=Font(name=AR,size=10); d[f"A{r}"].border=BORD
        d[f"D{r}"].value=cnt; d[f"D{r}"].font=Font(name=AR,size=10,bold=True); d[f"D{r}"].alignment=Alignment(horizontal="right"); d[f"D{r}"].border=BORD
    tr=16+len(locs); _nt(d,tr,1,"TOTAL DISCREPANCIES",4,tdisc,span=3)
    bg=BarChart(); bg.type="bar"; bg.title="Issues by location"; bg.height=8; bg.width=11; bg.style=11; bg.legend=None
    bg.add_data(Reference(d,min_col=4,min_row=16,max_row=16+len(locs)-1)); bg.set_categories(Reference(d,min_col=1,min_row=16,max_row=16+len(locs)-1))
    bl=DataLabelList(); bl.showVal=True; bl.showSerName=False; bg.dataLabels=bl
    try: bg.series[0].graphicalProperties.solidFill="C0504D"
    except: pass
    d.add_chart(bg,"E14")
    dr2=tr+2; m(f"A{dr2}:D{dr2}"); p(f"A{dr2}","Dollar impact",font=Font(name=AR,size=12,bold=True,color=NAVY),fill=SEC,align=Alignment(indent=1))
    m(f"A{dr2+1}:C{dr2+1}"); p(f"A{dr2+1}","Category",font=Font(name=AR,size=10,bold=True),fill="F2F2F2",border=True); p(f"D{dr2+1}","Per pay",font=Font(name=AR,size=10,bold=True),fill="F2F2F2",align=Alignment(horizontal="right"),border=True); p(f"E{dr2+1}","Per month",font=Font(name=AR,size=10,bold=True),fill="F2F2F2",align=Alignment(horizontal="right"),border=True)
    drows=[("Add - deductions to start",round(adds["Broker /pay"].sum(),2)),("Change - corrections needed",cdiff)]
    for i,(lab,amt) in enumerate(drows):
        r=dr2+2+i; m(f"A{r}:C{r}"); d[f"A{r}"].value=lab; d[f"A{r}"].font=Font(name=AR,size=10); d[f"A{r}"].border=BORD
        d[f"D{r}"].value=amt; d[f"D{r}"].font=Font(name=AR,size=10,bold=True); d[f"D{r}"].number_format="$#,##0.00"; d[f"D{r}"].alignment=Alignment(horizontal="right"); d[f"D{r}"].border=BORD
        d[f"E{r}"].value=round(amt*26/12,2); d[f"E{r}"].font=Font(name=AR,size=10,bold=True); d[f"E{r}"].number_format="$#,##0.00"; d[f"E{r}"].alignment=Alignment(horizontal="right"); d[f"E{r}"].border=BORD
    ta=round(sum(x[1] for x in drows),2); r=dr2+4; m(f"A{r}:C{r}"); _nt(d,r,1,"TOTAL",4,ta,fmt="$#,##0.00",span=3)
    d[f"E{r}"].value=round(ta*26/12,2); d[f"E{r}"].font=Font(name=AR,size=10,bold=True,color=WHITE); d[f"E{r}"].fill=PatternFill("solid",fgColor=TF); d[f"E{r}"].number_format="$#,##0.00"; d[f"E{r}"].alignment=Alignment(horizontal="right"); d[f"E{r}"].border=BORD
    d.page_setup.orientation="landscape"; d.page_setup.fitToWidth=1; d.page_setup.fitToHeight=0; d.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
    aw=wb.create_sheet("Action Items"); aw.sheet_view.showGridLines=False
    _sw(aw,{1:18,2:10,3:26,4:46,5:12,6:12,7:11,8:28,9:40})
    _nh(aw,"ACTION ITEMS  |  Color Key & Summary",9)
    _kh(aw,[("Color / Action",1,1),("Count",2,2),("What it means",3,5),("What to do",6,9)])
    aik=[("DDEBF7","Blue  -  Add",ct["Add"],"Enrolled with broker but zero deduction in Paycor. Company may be eating this cost.","Set up the deduction in Paycor immediately."),
         ("FCE4D6","Pink  -  Change",ct["Change"],"Both sides have the benefit but dollar amounts differ.","Update the deduction amount in Paycor to match the broker."),
         ("FFF2CC","Yellow  -  Review",ct["Review"],"Benefit not in Paycor export, or employee name did not match.","Verify manually in Paycor.")]
    for ri,(fc,lb,cnt,mn,ac) in enumerate(aik,3): _kr(aw,ri,fc,lb,cnt,mn,ac,5,6,9)
    _kt(aw,6,"TOTAL ACTION ITEMS",tdisc,f"{ct['Add']} Add  +  {ct['Change']} Change  +  {ct['Review']} Review  =  {tdisc} lines",f"{n_emps} employees need attention",3,5,6,9)
    aw.row_dimensions[7].height=8; _dh(aw,DISPLAY_COLS,8); _dr(aw,disc,9); aw.column_dimensions["J"].width=40
    allw=wb.create_sheet("All Lines"); allw.sheet_view.showGridLines=False
    _sw(allw,{1:18,2:13,3:13,4:11,5:28,6:12,7:12,8:11,9:26,10:40})
    _nh(allw,"ALL RECONCILIATION LINES  |  Color Key & Summary",10)
    _kh(allw,[("Color / Action",1,1),("Count",2,2),("What it means",3,6),("What to do",7,10)])
    allk=[("E2EFDA","Green  -  OK",ct["OK"],"Broker and Paycor match perfectly.","No action needed."),
          ("DDEBF7","Blue  -  Add",ct["Add"],"Enrolled with broker but zero deduction in Paycor.","Set up the deduction in Paycor immediately."),
          ("FCE4D6","Pink  -  Change",ct["Change"],"Both sides have the benefit but dollar amounts differ.","Update the deduction amount in Paycor to match the broker."),
          ("FFF2CC","Yellow  -  Review",ct["Review"],"Benefit not in export or name not matched.","Verify manually in Paycor.")]
    for ri,(fc,lb,cnt,mn,ac) in enumerate(allk,3): _kr(allw,ri,fc,lb,cnt,mn,ac,6,7,10)
    _kt(allw,7,"TOTAL LINES",total,f"{ct['OK']} OK  +  {ct['Add']} Add  +  {ct['Change']} Change  +  {ct['Review']} Review  =  {total}",f"{tdisc} lines need attention  |  {n_emps} employees affected",3,6,7,10)
    allw.row_dimensions[8].height=8; _dh(allw,DISPLAY_COLS,9); _dr(allw,df,10)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ── HISTORY (Supabase) ────────────────────────────────────────────
def save_to_history(recon_df, client_name, period, excel_bytes=None):
    ct = {
        "OK":     int((recon_df["Action"] == "OK").sum()),
        "Add":    int(recon_df["Action"].str.startswith("Add").sum()),
        "Change": int(recon_df["Action"].str.startswith("Change").sum()),
        "Review": int(recon_df["Action"].str.startswith("Review").sum()),
    }
    mo = round(recon_df[recon_df["Action"].str.startswith("Add")]["Broker /pay"].sum() * 26 / 12, 2)
    report_file = ""
    if excel_bytes:
        safe = client_name.strip().replace(" ", "_").replace("/", "_")
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = os.path.join(REPORTS_DIR, f"PayMatch_{safe}_{period.replace(' ','_')}_{ts}.xlsx")
        with open(report_file, "wb") as f:
            f.write(excel_bytes)
    try:
        existing = _sb().table("reconciliation_history").select("id,notes").eq("client", client_name).eq("period", period).execute()
        record = ReconciliationRecord(
            client=client_name, period=period, run_by="",
            run_date=datetime.now().strftime("%Y-%m-%d"),
            run_time=datetime.now().strftime("%I:%M %p"),
            total_lines=len(recon_df),
            ok_count=ct["OK"], add_count=ct["Add"],
            change_count=ct["Change"], review_count=ct["Review"],
            discrepancies=ct["Add"] + ct["Change"] + ct["Review"],
            monthly_at_stake=mo, report_file=report_file, notes="",
        )
        row = record.model_dump()
        if existing.data:
            row["notes"] = existing.data[0].get("notes", "") or ""
            _sb().table("reconciliation_history").update(row).eq("client", client_name).eq("period", period).execute()
        else:
            _sb().table("reconciliation_history").insert(row).execute()
    except Exception:
        pass

def load_history(client_name=None):
    try:
        q = _sb().table("reconciliation_history").select("*").order("run_date", desc=True)
        if client_name:
            q = q.eq("client", client_name)
        res = q.execute()
        if not res.data:
            return pd.DataFrame()
        return pd.DataFrame(res.data).rename(columns={
            "ok_count": "OK", "add_count": "Add", "change_count": "Change",
            "review_count": "Review", "total_lines": "Total Lines",
            "discrepancies": "Discrepancies",
            "monthly_at_stake": "Monthly $ at stake", "run_by": "Run By",
            "run_date": "Run Date", "run_time": "Run Time",
            "report_file": "Report File", "client": "Client",
            "period": "Period", "notes": "Notes", "status": "Status",
        }).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

def get_run_notes(client, period):
    try:
        res = _sb().table("reconciliation_history").select("notes").eq("client", client).eq("period", period).execute()
        return (res.data[0].get("notes", "") or "") if res.data else ""
    except Exception:
        return ""

def update_run_notes(client, period, notes):
    try:
        _sb().table("reconciliation_history").update({"notes": str(notes)}).eq("client", client).eq("period", period).execute()
    except Exception:
        pass

# ── LOGO (base64 for embedding in HTML header) ────────────────────
@st.cache_resource
def _logo_b64() -> str:
    logo_path = os.path.join(_APP_DIR, "ProPayHR-2024-red-TRANSPARENT.png")
    if not os.path.exists(logo_path):
        return ""
    with open(logo_path, "rb") as f:
        return base64.b64encode(f.read()).decode()

# ─────────────────────────────────────────────────────────────────
# STREAMLIT APP
# ─────────────────────────────────────────────────────────────────
st.set_page_config(page_title="PayMatch by ProPayHR", page_icon="📊", layout="wide")

# ── KEEP-ALIVE ────────────────────────────────────────────────────
@st.cache_resource
def _start_keep_alive():
    def _loop():
        while True:
            time.sleep(60)
    t = threading.Thread(target=_loop, daemon=True)
    t.start()
    return t

_start_keep_alive()

st.markdown(f"<style>{APP_CSS}</style>", unsafe_allow_html=True)

# ── SESSION STATE ─────────────────────────────────────────────────
_defaults = {
    "recon_df":           None,
    "last_run_ts":        None,
    "delete_confirm_key": None,
}
for _k, _v in _defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── HEADER ────────────────────────────────────────────────────────
_logo = _logo_b64()
_logo_html = (
    f'<img src="data:image/png;base64,{_logo}" '
    f'style="height:44px;width:auto;object-fit:contain;display:block;margin-bottom:0.85rem;" alt="ProPayHR" />'
    if _logo else ""
)

st.markdown(f"""
<div class="pm-hero">
    <div class="pm-hero-inner">
        <div class="pm-hero-left">
            {_logo_html}
            <div class="pm-superlabel">ProPayHR &nbsp;·&nbsp; Enterprise Benefits Platform</div>
            <div class="pm-wordmark">Pay<em>Match</em></div>
            <div class="pm-descriptor">Automated Benefits Reconciliation &nbsp;&nbsp;·&nbsp;&nbsp; Broker vs Paycor &nbsp;&nbsp;·&nbsp;&nbsp; Payroll Accuracy</div>
            <div class="pm-chips">
                <span class="pm-chip"><span class="pm-chip-check">✓</span>Broker vs Paycor Comparison</span>
                <span class="pm-chip"><span class="pm-chip-check">✓</span>Any Client, Any Month</span>
                <span class="pm-chip"><span class="pm-chip-check">✓</span>Automatic Benefit Matching</span>
                <span class="pm-chip"><span class="pm-chip-check">✓</span>Excel Report &amp; Dashboard Export</span>
                <span class="pm-chip"><span class="pm-chip-check">✓</span>Run History &amp; Trend View</span>
            </div>
        </div>
    </div>
</div>
<div class="gold-rule"></div>
""", unsafe_allow_html=True)

# ── UPLOAD + CONFIG ───────────────────────────────────────────────
left, right = st.columns([3, 2], gap="large")

with left:
    st.markdown('<div class="slabel">Upload Files</div>', unsafe_allow_html=True)
    st.markdown('<div class="sdesc">Upload both files below, then click Run. PayMatch compares them line by line and flags every mismatch.</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="card pm-step-card">
        <div class="card-title">
            <span class="step-badge">1</span>
            Deduction Report
            <span class="file-tag">Paycor</span>
        </div>
        <div class="card-desc">Export from Paycor — what Paycor is currently deducting per paycheck. Accepted: .xlsx, .xls, .csv</div>
    </div>
    """, unsafe_allow_html=True)
    paycor_file = st.file_uploader("Deduction Report", type=["xlsx","xls","csv"],
                                    label_visibility="collapsed", key="paycor")
    if paycor_file and not paycor_file.name.lower().endswith((".xlsx",".xls",".csv")):
        st.error(f"'{paycor_file.name}' isn't a supported file type. Please upload an Excel (.xlsx / .xls) or CSV file.")

    st.markdown("<div style='height:0.4rem'></div>", unsafe_allow_html=True)

    st.markdown("""
    <div class="card pm-step-card">
        <div class="card-title">
            <span class="step-badge">2</span>
            Master Mapping Report
            <span class="file-tag">Broker</span>
        </div>
        <div class="card-desc">Enrollment export from your broker — one row per person per benefit, with EE Cost per pay. PayMatch reads the benefit names automatically. Accepted: .xlsx, .xls, .csv</div>
    </div>
    """, unsafe_allow_html=True)
    broker_file = st.file_uploader("Master Mapping Report", type=["xlsx","xls","csv"],
                                    label_visibility="collapsed", key="broker")
    if broker_file and not broker_file.name.lower().endswith((".xlsx",".xls",".csv")):
        st.error(f"'{broker_file.name}' isn't a supported file type. Please upload an Excel (.xlsx / .xls) or CSV file.")

with right:
    st.markdown('<div class="slabel">Configuration</div>', unsafe_allow_html=True)
    st.markdown('<div class="sdesc">Enter the client name and choose the month you\'re reconciling. This labels your report and saves the run to history.</div>', unsafe_allow_html=True)

    client_name = st.text_input("Client Name", value="", placeholder="e.g. Acme Corporation",
        help="The name of the client this reconciliation is for. It will appear in the report header.")

    mo_col, yr_col = st.columns(2)
    with mo_col:
        sel_month = st.selectbox("Month", options=MONTHS,
            index=datetime.now().month - 1, key="period_month")
    with yr_col:
        sel_year = st.number_input("Year", min_value=2020, max_value=2040,
            value=datetime.now().year, step=1, key="period_year", format="%d")
    period = f"{sel_month} {int(sel_year)}"

    tolerance = st.number_input("Acceptable Difference ($)", min_value=0.0, max_value=1.0,
        value=0.01, step=0.01,
        help="Two amounts this close to each other will be treated as a match.")
    st.markdown('<div style="font-size:0.78rem;color:#7D5A5E;line-height:1.55;margin-top:0.3rem;margin-bottom:0.6rem;">How close two dollar amounts need to be to count as a match. The default ($0.01) treats anything more than a penny apart as a discrepancy.</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box">
        <b>How PayMatch works</b>
        Both files use per-paycheck dollar amounts and are compared directly. Benefit labels are read automatically from your enrollment file — nothing needs to be configured manually.
        Results: <strong>OK</strong> = match &nbsp;·&nbsp; <strong>Add</strong> = missing deduction &nbsp;·&nbsp; <strong>Change</strong> = amount differs &nbsp;·&nbsp; <strong>Review</strong> = needs manual check.
    </div>
    """, unsafe_allow_html=True)

st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
st.divider()

# ── RUN BUTTON ────────────────────────────────────────────────────
ready = paycor_file is not None and broker_file is not None and bool(client_name.strip())
_, btn_col, _ = st.columns([1, 2, 1])
with btn_col:
    run = st.button("Run Reconciliation", type="primary", disabled=not ready)

if not ready and not run:
    missing = []
    if paycor_file is None: missing.append("Deduction Report")
    if broker_file is None: missing.append("Master Mapping Report")
    if not client_name.strip(): missing.append("Client Name")
    st.markdown(f'<div class="empty-prompt">Please provide your <strong>{", ".join(missing)}</strong> above to get started.</div>', unsafe_allow_html=True)

# ── RESULTS ───────────────────────────────────────────────────────
if run:
    with st.spinner("Analyzing files and comparing broker enrollment against Paycor deductions…"):
        try:
            paycor_df            = parse_paycor(paycor_file)
            broker_df, crosswalk = parse_broker_and_crosswalk(broker_file)
            if len(paycor_df) == 0:
                st.error(f"The Deduction Report ({paycor_file.name}) appears to be empty or couldn't be read. Make sure it's a valid Paycor export with at least one row of data.")
                st.stop()
            if len(broker_df) == 0:
                st.error(f"The Master Mapping Report ({broker_file.name}) appears to be empty or couldn't be read. Make sure the file has benefit labels in the second row and employee data below that.")
                st.stop()
            recon_df = reconcile(broker_df, paycor_df, crosswalk, tolerance)
            st.session_state.recon_df    = recon_df
            st.session_state.last_run_ts = datetime.now()
            _excel_full    = build_excel(recon_df, client_name.strip(), period=period)
            _excel_actions = build_action_items_only(recon_df, client_name.strip())
            save_to_history(recon_df, client_name.strip(), period, excel_bytes=_excel_full)

            if len(recon_df) == 0:
                st.warning("No matching employees found. Make sure first and last names are spelled the same way in both files, and that the enrollment file includes per-paycheck dollar amounts.")
            else:
                total  = len(recon_df)
                n_ok   = int((recon_df["Action"]=="OK").sum())
                n_add  = int(recon_df["Action"].str.startswith("Add").sum())
                n_chg  = int(recon_df["Action"].str.startswith("Change").sum())
                n_rev  = int(recon_df["Action"].str.startswith("Review").sum())
                add_pp = round(recon_df[recon_df["Action"].str.startswith("Add")]["Broker /pay"].sum(), 2)
                _nf = recon_df[(recon_df["Action"] == "Review - Not in Paycor") & (recon_df["Broker /pay"] > 0)]
                notfound_pp = round(_nf["Broker /pay"].sum(), 2)
                mo_var = round((add_pp + notfound_pp)*26/12, 2)
                n_emps = recon_df[recon_df["Action"]!="OK"]["First Name"].nunique()
                disc_ct = n_add + n_chg + n_rev
                ts = st.session_state.last_run_ts.strftime("%-I:%M %p on %B %-d, %Y")

                st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)

                if disc_ct == 0:
                    st.success(f"All {total:,} benefit lines match — nothing needs attention.")
                else:
                    st.markdown(f"""
                    <div class="res-bar">
                        <div>
                            <div class="res-bar-ttl">Reconciliation Complete &nbsp;·&nbsp; {client_name.strip()} &nbsp;·&nbsp; {period}</div>
                            <div class="res-bar-sub">{total:,} lines compared &nbsp;·&nbsp; {n_emps} employees with discrepancies</div>
                            <div class="run-ts">&#9201; Ran at {ts}</div>
                        </div>
                        <div class="res-bar-count">{disc_ct} need attention</div>
                    </div>
                    """, unsafe_allow_html=True)

                st.markdown(f"""
                <div class="pm-metrics">
                    <div class="pm-metric">
                        <div class="pm-metric-label">Lines Compared</div>
                        <div class="pm-metric-value"><span class="pm-countup">{total:,}</span></div>
                    </div>
                    <div class="pm-metric">
                        <div class="pm-metric-label">Matched OK</div>
                        <div class="pm-metric-value"><span class="pm-countup">{n_ok:,}</span></div>
                    </div>
                    <div class="pm-metric">
                        <div class="pm-metric-label">Add to Paycor</div>
                        <div class="pm-metric-value"><span class="pm-countup">{n_add:,}</span></div>
                    </div>
                    <div class="pm-metric">
                        <div class="pm-metric-label">Amount Change</div>
                        <div class="pm-metric-value"><span class="pm-countup">{n_chg:,}</span></div>
                    </div>
                    <div class="pm-metric">
                        <div class="pm-metric-label">Needs Review</div>
                        <div class="pm-metric-value"><span class="pm-countup">{n_rev:,}</span></div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if mo_var > 0:
                    st.markdown(f"""
                    <div class="urgency">
                        <div style="font-size:1.15rem;flex-shrink:0;margin-top:0.05rem;">&#9888;</div>
                        <div>
                            <div class="urgency-amount"><span class="pm-countup">${mo_var:,.2f}</span> / month not being collected</div>
                            <div class="urgency-note">Deductions missing for matched employees: ${add_pp:,.2f} per pay &nbsp;·&nbsp; Enrolled employees not found in payroll: ${notfound_pp:,.2f} per pay. These need to be corrected immediately to avoid further financial exposure.</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                # Status legend
                st.markdown("""
                <div style="display:flex;align-items:center;gap:0.5rem;margin:1.1rem 0 0.5rem;flex-wrap:wrap;">
                    <span style="font-size:0.65rem;font-weight:700;letter-spacing:0.14em;text-transform:uppercase;color:#9D8E88;margin-right:0.25rem;">Status key</span>
                    <span class="badge badge-ok">&#10003; OK</span>
                    <span class="badge badge-add">+ Add</span>
                    <span class="badge badge-change">&#8645; Change</span>
                    <span class="badge badge-review">&#9679; Review</span>
                </div>
                """, unsafe_allow_html=True)

                disc = recon_df[recon_df["Action"]!="OK"]
                if len(disc) > 0:
                    st.markdown(f"""
                    <div style="display:flex;align-items:center;justify-content:space-between;margin:0.5rem 0 0.5rem;flex-wrap:wrap;gap:0.5rem;">
                        <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.16em;text-transform:uppercase;color:#9D8E88;">Action Items — what needs fixing</div>
                        <div style="background:rgba(139,26,47,0.1);color:#8B1A2F;border:1px solid rgba(139,26,47,0.2);border-radius:999px;padding:0.18rem 0.75rem;font-size:0.72rem;font-weight:700;">{len(disc)} lines &nbsp;·&nbsp; {n_emps} employees</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.dataframe(disc[DISPLAY_COLS], use_container_width=True, hide_index=True)

                # Downloads
                st.markdown("""
                <div class="dl-section">
                    <div class="dl-title">Download Results</div>
                </div>
                """, unsafe_allow_html=True)
                dl1, dl2 = st.columns(2)
                safe_name = client_name.strip().replace(' ', '_')
                date_str  = datetime.now().strftime('%Y%m%d')
                with dl1:
                    st.markdown('<div class="dl-desc"><strong>Full Excel Report</strong><br>Dashboard, all action items, and every reconciliation line with color coding.</div>', unsafe_allow_html=True)
                    st.download_button(
                        "Download Full Report (.xlsx)",
                        data=_excel_full,
                        file_name=f"PayMatch_{safe_name}_{date_str}_Full.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_full")
                with dl2:
                    st.markdown('<div class="dl-desc"><strong>Action Items Only</strong><br>Just the discrepancies — a clean list to hand off for corrections.</div>', unsafe_allow_html=True)
                    st.download_button(
                        "Download Action Items (.xlsx)",
                        data=_excel_actions,
                        file_name=f"PayMatch_{safe_name}_{date_str}_ActionItems.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_actions")

        except Exception as e:
            err_str = str(e).lower()
            if "no sheet" in err_str or "worksheet" in err_str:
                st.error("Couldn't read one of the files — it may be password-protected or have an unexpected layout. Try exporting a fresh copy from Paycor or your broker.")
            elif "codec" in err_str or "decode" in err_str or "encoding" in err_str:
                st.error("One of the files couldn't be opened correctly. If it's a CSV, try opening it in Excel and saving again before uploading.")
            elif "column" in err_str or "key" in err_str:
                st.error("A required column couldn't be found. Make sure the Deduction Report is a standard Paycor export and the enrollment file has employee names in columns labeled 'First Name' and 'Last Name'.")
            else:
                st.error(f"Something went wrong while processing: {e}")
            st.info("If the problem persists, double-check that the Deduction Report is from Paycor and the Master Mapping Report is your broker's enrollment file with benefit labels in the second row.")

# ── TREND VIEW (per-client, shown when a client name is entered) ──
if client_name.strip():
    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
    history_df = load_history(client_name.strip())

    if len(history_df) > 1:
        st.markdown('<div class="trend-card">', unsafe_allow_html=True)
        st.markdown('<div class="slabel">Month-over-Month Trend</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="sdesc">How discrepancies and uncollected dollars have changed across reconciliation runs for <strong>{client_name.strip()}</strong>.</div>', unsafe_allow_html=True)

        t1, t2 = st.columns(2)
        with t1:
            st.markdown("**Discrepancies by month**")
            chart_data = history_df.sort_values("Period").set_index("Period")[["Add","Change","Review"]]
            st.bar_chart(chart_data, color=["#8B1A2F","#C9A227","#B23A22"])
        with t2:
            st.markdown("**Monthly $ at stake**")
            mo_data = history_df.sort_values("Period").set_index("Period")[["Monthly $ at stake"]]
            st.line_chart(mo_data, color=["#8B1A2F"])

        st.markdown("**Run history**")
        dh = history_df[["Period","Run Date","Total Lines","OK","Add","Change","Review","Monthly $ at stake"]].copy()
        dh["Monthly $ at stake"] = dh["Monthly $ at stake"].apply(lambda x: f"${x:,.2f}")
        st.dataframe(dh, use_container_width=True, hide_index=True)

        if len(history_df) >= 2:
            latest = history_df.iloc[0]; prev = history_df.iloc[1]
            dc = int(latest["Discrepancies"]) - int(prev["Discrepancies"])
            mc = round(float(latest["Monthly $ at stake"]) - float(prev["Monthly $ at stake"]), 2)
            cls = "delta-dn" if dc < 0 else "delta-up"
            st.markdown(f"""
            <div class="delta {cls}">
                Discrepancies {"&#8595;" if dc<0 else "&#8593;"} {abs(dc)} vs previous run
                &nbsp;·&nbsp; Monthly $ {"&#8595;" if mc<0 else "&#8593;"} ${abs(mc):,.2f}
            </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    elif len(history_df) == 1:
        st.markdown(f'<div style="background:var(--paper,white);border-radius:12px;padding:1rem 1.5rem;border:1px solid #E3E0D8;color:#9D8E88;font-size:0.84rem;margin-top:0.5rem;">Trend view will appear after your second reconciliation run for <strong style="color:#5C544E">{client_name.strip()}</strong>.</div>', unsafe_allow_html=True)

    # ── NOTES ─────────────────────────────────────────────────────
    if len(history_df) > 0:
        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        st.markdown('<div class="notes-wrap">', unsafe_allow_html=True)
        st.markdown('<div class="notes-label">Run Notes</div>', unsafe_allow_html=True)
        st.markdown(f'<div style="font-size:0.82rem;color:#7D5A5E;margin-bottom:0.8rem;">Notes for <strong style="color:#2A0E14;">{client_name.strip()}</strong> &nbsp;·&nbsp; {period}. Saved automatically with this run.</div>', unsafe_allow_html=True)
        _nk = f"rn_{''.join(c if c.isalnum() else '_' for c in client_name.strip())}_{period.replace(' ','_')}"
        if _nk not in st.session_state:
            st.session_state[_nk] = get_run_notes(client_name.strip(), period)
        st.text_area("Notes", key=_nk, placeholder="e.g. Sent corrections to payroll team on 6/10. 3 employees pending broker re-enrollment.", label_visibility="collapsed", height=100)
        if st.button("Save Notes", key="save_run_notes"):
            update_run_notes(client_name.strip(), period, st.session_state[_nk])
            st.success("Notes saved.")
        st.markdown('</div>', unsafe_allow_html=True)

# ── FOOTER ────────────────────────────────────────────────────────
st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)
st.markdown(f"""
<div class="pm-footer">
    <strong>PayMatch</strong> by ProPayHR &nbsp;·&nbsp; Benefits Reconciliation Platform &nbsp;·&nbsp; {datetime.now().year}
</div>
<script>
(function(){{
    if(!window.__pmAlive){{
        window.__pmAlive = true;
        setInterval(function(){{
            try{{fetch(window.location.origin + window.location.pathname, {{method:'HEAD', mode:'no-cors', cache:'no-store'}})}}catch(e){{}}
        }}, 45000);
    }}
}})();
</script>
""", unsafe_allow_html=True)
